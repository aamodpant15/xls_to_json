#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json
import openpyxl as xl
from sys import argv


aIndex= argv.index("-a")

fileName= argv[1]

sheetName= ""
for args in range(aIndex+1,len(argv)):
    sheetName = sheetName + " "+ argv[args]
sheetName = sheetName.strip()

if(len(argv) <3):
    print ("\nUsage: ./json_from_xlsx <Filename.xlsx> -a <sheet_name>\n")

#open workbook
workbook = xl.load_workbook(fileName)

#get sheet
sheet = workbook[sheetName]

#member count
memCount= 0

#number of rows
X = sheet.max_row

#number of columns
Y = sheet.max_column

#column titles
attributes= []

for col in range (1, Y+1):
    title = sheet.cell(row = 1, column= col).value
    attributes.append(title)

#members
membas= [] 

#keep count of which attribute from the array is being added
attributeCount= 0

# print(len(attributes))
idCount=1

for rows in range(2, X+1):
    attributeCount= 0
    membas.append({})
    memba = membas[memCount]
    memba["id"] = idCount
    idCount+=1
    for cols in range(1, Y+1):
        attribute= attributes[attributeCount]
        data = str(sheet.cell(row=rows, column=cols).value)
        memba[attribute]=data
        attributeCount+= 1
    memCount+= 1



json.dump(membas, open("result.json", "w"))
